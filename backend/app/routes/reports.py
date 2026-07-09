"""
FalconOps AI - Reports Routes
Report generation, scheduling, and export
"""
import uuid
import io
from datetime import datetime, timezone, timedelta
from typing import Optional, List
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse

from ..core.database import db
from ..models.schemas import (
    ScheduledReportCreate, ScheduledReportResponse, 
    ExecutiveReportResponse, ReportDateRange
)
from ..utils.auth import require_auth, require_write_access, get_current_user
from ..services.reports_service import (
    generate_uptime_report,
    generate_executive_report_data,
    generate_sla_report_data,
    generate_incident_report_data,
    generate_report_pdf,
    send_scheduled_report
)

router = APIRouter(prefix="/api/reports", tags=["Reports"])


def _tid(user):
    return user.get("tenant_id") if user else None


def get_default_dates():
    """Get default date range (last 7 days)"""
    end = datetime.now(timezone.utc)
    start = end - timedelta(days=7)
    return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")


@router.get("/uptime")
async def get_uptime_report(
    hours: int = Query(24, ge=1, le=720),
    monitor_ids: Optional[str] = Query(None, description="Comma-separated monitor IDs"),
    current_user: Optional[dict] = Depends(get_current_user)
):
    """Get uptime report for specified period"""
    ids = monitor_ids.split(",") if monitor_ids else None
    return await generate_uptime_report(period_hours=hours, monitor_ids=ids)


@router.get("/executive", response_model=ExecutiveReportResponse)
async def get_executive_report(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    include_ai_summary: bool = Query(True),
    current_user: Optional[dict] = Depends(get_current_user)
):
    """Get executive report with KPIs and AI summary"""
    if not start_date or not end_date:
        start_date, end_date = get_default_dates()
    
    report_data = await generate_executive_report_data(start_date, end_date, include_ai_summary)
    return ExecutiveReportResponse(
        period=report_data["period"],
        kpis=report_data["kpis"],
        incident_trends=[],
        category_breakdown=report_data.get("category_breakdown", []),
        team_performance=[],
        sla_summary=report_data["sla_summary"],
        ai_summary=report_data.get("ai_summary"),
        generated_at=report_data["generated_at"]
    )


@router.get("/sla")
async def get_sla_report(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    current_user: Optional[dict] = Depends(get_current_user)
):
    """Get SLA compliance report"""
    if not start_date or not end_date:
        start_date, end_date = get_default_dates()
    return await generate_sla_report_data(start_date, end_date)


@router.get("/incidents")
async def get_incidents_report(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    current_user: Optional[dict] = Depends(get_current_user)
):
    """Get incident analytics report"""
    if not start_date or not end_date:
        start_date, end_date = get_default_dates()
    return await generate_incident_report_data(start_date, end_date)


@router.get("/team-performance")
async def get_team_performance_report(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    current_user: Optional[dict] = Depends(get_current_user)
):
    """Get team performance metrics"""
    if not start_date or not end_date:
        start_date, end_date = get_default_dates()
    
    incidents = await db.incidents.find({}, {"_id": 0}).to_list(1000)
    
    resolved = [i for i in incidents if i.get("status") == "resolved"]
    mttr_values = [i.get("mttr_seconds", 0) for i in resolved if i.get("mttr_seconds")]
    
    return {
        "period": {"start_date": start_date, "end_date": end_date},
        "metrics": {
            "total_incidents_handled": len(incidents),
            "incidents_resolved": len(resolved),
            "avg_mttr_minutes": round(sum(mttr_values) / len(mttr_values) / 60, 1) if mttr_values else 0,
            "resolution_rate": round(len(resolved) / len(incidents) * 100, 1) if incidents else 100
        },
        "generated_at": datetime.now(timezone.utc).isoformat()
    }


@router.get("/export/pdf")
async def export_report_pdf(
    report_type: str = Query("executive"),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    current_user: dict = Depends(require_auth)
):
    """Export report as PDF"""
    if not start_date or not end_date:
        start_date, end_date = get_default_dates()
    
    if report_type == "executive":
        report_data = await generate_executive_report_data(start_date, end_date)
    elif report_type == "sla":
        report_data = await generate_sla_report_data(start_date, end_date)
    elif report_type == "incidents":
        report_data = await generate_incident_report_data(start_date, end_date)
    else:
        report_data = await generate_uptime_report(period_hours=168)
    
    pdf_bytes = await generate_report_pdf(report_type, report_data, "Custom Period")
    
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=falconops_{report_type}_report.pdf"}
    )


@router.get("/export/excel")
async def export_report_excel(
    report_type: str = Query("executive"),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    current_user: dict = Depends(require_auth)
):
    """Export report as Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    if not start_date or not end_date:
        start_date, end_date = get_default_dates()
    
    if report_type == "executive":
        report_data = await generate_executive_report_data(start_date, end_date)
    elif report_type == "sla":
        report_data = await generate_sla_report_data(start_date, end_date)
    elif report_type == "incidents":
        report_data = await generate_incident_report_data(start_date, end_date)
    else:
        report_data = await generate_uptime_report(period_hours=168)
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"{report_type.title()} Report"
    
    ws["A1"] = "FalconOps AI Report"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"Period: {start_date} to {end_date}"
    ws["A3"] = f"Generated: {datetime.now(timezone.utc).isoformat()}"
    
    row = 5
    if "kpis" in report_data:
        ws[f"A{row}"] = "Key Performance Indicators"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1
        for key, value in report_data["kpis"].items():
            ws[f"A{row}"] = key.replace("_", " ").title()
            ws[f"B{row}"] = str(value)
            row += 1
    
    if "sla_summary" in report_data or "summary" in report_data:
        row += 1
        ws[f"A{row}"] = "SLA Summary"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1
        summary = report_data.get("sla_summary", report_data.get("summary", {}))
        for key, value in summary.items():
            ws[f"A{row}"] = key.replace("_", " ").title()
            ws[f"B{row}"] = str(value)
            row += 1
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=falconops_{report_type}_report.xlsx"}
    )


# Scheduled Reports CRUD

@router.get("/scheduled", response_model=List[ScheduledReportResponse])
async def get_scheduled_reports(current_user: dict = Depends(require_auth)):
    """Get all scheduled reports"""
    reports = await db.scheduled_reports.find({}, {"_id": 0}).to_list(100)
    return [ScheduledReportResponse(**r) for r in reports]


@router.post("/scheduled", response_model=ScheduledReportResponse)
async def create_scheduled_report(report: ScheduledReportCreate, current_user: dict = Depends(require_write_access)):
    """Create a scheduled report"""
    report_id = str(uuid.uuid4())
    report_doc = {
        "id": report_id,
        "name": report.name,
        "frequency": report.frequency,
        "recipients": report.recipients,
        "report_type": report.report_type,
        "day_of_week": report.day_of_week,
        "hour": report.hour,
        "enabled": report.enabled,
        "include_monitors": report.include_monitors,
        "include_pdf": report.include_pdf,
        "include_ai_summary": report.include_ai_summary,
        "last_sent": None,
        "next_run": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user["email"]
    }
    
    await db.scheduled_reports.insert_one(report_doc)
    return ScheduledReportResponse(**report_doc)


@router.put("/scheduled/{report_id}", response_model=ScheduledReportResponse)
async def update_scheduled_report(report_id: str, report: ScheduledReportCreate, current_user: dict = Depends(require_write_access)):
    """Update a scheduled report"""
    existing = await db.scheduled_reports.find_one({"id": report_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Scheduled report not found")
    
    update_data = {
        "name": report.name,
        "frequency": report.frequency,
        "recipients": report.recipients,
        "report_type": report.report_type,
        "day_of_week": report.day_of_week,
        "hour": report.hour,
        "enabled": report.enabled,
        "include_monitors": report.include_monitors,
        "include_pdf": report.include_pdf,
        "include_ai_summary": report.include_ai_summary,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.scheduled_reports.update_one({"id": report_id}, {"$set": update_data})
    updated = await db.scheduled_reports.find_one({"id": report_id}, {"_id": 0})
    return ScheduledReportResponse(**updated)


@router.delete("/scheduled/{report_id}")
async def delete_scheduled_report(report_id: str, current_user: dict = Depends(require_write_access)):
    """Delete a scheduled report"""
    result = await db.scheduled_reports.delete_one({"id": report_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Scheduled report not found")
    return {"message": "Scheduled report deleted"}


@router.post("/scheduled/{report_id}/send")
async def send_report_now(report_id: str, current_user: dict = Depends(require_write_access)):
    """Manually trigger a scheduled report"""
    report = await db.scheduled_reports.find_one({"id": report_id}, {"_id": 0})
    if not report:
        raise HTTPException(status_code=404, detail="Scheduled report not found")
    
    success = await send_scheduled_report(report)
    
    if success:
        return {"message": "Report sent successfully"}
    else:
        raise HTTPException(status_code=500, detail="Failed to send report")
