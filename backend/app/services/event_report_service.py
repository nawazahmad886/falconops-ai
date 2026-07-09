"""
FalconOps AI - Event Report Export Service
Enterprise-grade Excel & PDF report generation for event analysis
"""
import io
import os
import uuid
from datetime import datetime, timezone
from typing import Dict, Any, List, Optional

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart as XlPieChart, BarChart as XlBarChart, Reference
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch, mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    Image as RLImage, PageBreak, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT


# ── colour constants ────────────────────────────────────────
SEV_COLORS = {
    "critical": "EF4444", "warning": "F59E0B", "info": "06B6D4",
}
SEV_FILLS = {k: PatternFill(start_color=v, end_color=v, fill_type="solid") for k, v in SEV_COLORS.items()}
HEADER_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"),
)

# ReportLab colours
RL_SEV = {"critical": colors.HexColor("#EF4444"), "warning": colors.HexColor("#F59E0B"), "info": colors.HexColor("#06B6D4")}
RL_DARK = colors.HexColor("#0F172A")
RL_ACCENT = colors.HexColor("#00E0FF")


# ═══════════════════════════════════════════════════════════
#  EXCEL GENERATION
# ═══════════════════════════════════════════════════════════

def generate_excel_report(
    analysis_result: Dict[str, Any],
    events: List[Dict],
    branding: Optional[Dict] = None,
) -> io.BytesIO:
    """Generate multi-sheet Excel workbook."""

    wb = Workbook()
    brand = branding or {}
    report_title = brand.get("title", "FalconOps AI - Event Analysis Report")

    # ── Sheet 1: Executive Summary ──────────────────────
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.sheet_properties.tabColor = "00E0FF"

    summary = analysis_result.get("summary", {})
    patterns = analysis_result.get("patterns", {})

    # Title
    ws1.merge_cells("A1:F1")
    ws1["A1"] = report_title
    ws1["A1"].font = Font(bold=True, size=16, color="0F172A")
    ws1["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws1["A2"].font = Font(italic=True, color="6B7280", size=10)
    if brand.get("company"):
        ws1["A3"] = f"Prepared for: {brand['company']}"
        ws1["A3"].font = Font(italic=True, color="6B7280", size=10)

    # KPI table
    row = 5
    kpi_headers = ["Metric", "Value"]
    for ci, h in enumerate(kpi_headers, 1):
        c = ws1.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER

    sev = patterns.get("severity_distribution", {})
    total = summary.get("total_events", 0)
    critical = sev.get("critical", 0)
    warning = sev.get("warning", 0)
    info = sev.get("info", 0)

    kpis = [
        ("Total Events", total),
        ("Critical Alerts", critical),
        ("Warning Alerts", warning),
        ("Info Alerts", info),
        ("Health Score", f"{summary.get('health_score', 0)}%"),
        ("Services Affected", summary.get("services_affected", 0)),
        ("Hosts Affected", summary.get("hosts_affected", 0)),
        ("Unique Alert Types", patterns.get("unique_alerts", 0)),
    ]
    for i, (k, v) in enumerate(kpis, row + 1):
        ws1.cell(row=i, column=1, value=k).border = THIN_BORDER
        vc = ws1.cell(row=i, column=2, value=v)
        vc.border = THIN_BORDER
        vc.font = Font(bold=True)

    # Severity distribution mini-table
    sev_row = row + len(kpis) + 2
    ws1.cell(row=sev_row, column=1, value="Severity Distribution").font = Font(bold=True, size=12)
    sev_row += 1
    for ci, h in enumerate(["Severity", "Count", "Percentage"], 1):
        c = ws1.cell(row=sev_row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
    for sname, scount in sev.items():
        sev_row += 1
        ws1.cell(row=sev_row, column=1, value=sname.upper()).border = THIN_BORDER
        ws1.cell(row=sev_row, column=2, value=scount).border = THIN_BORDER
        pct = round(scount / total * 100, 1) if total else 0
        ws1.cell(row=sev_row, column=3, value=f"{pct}%").border = THIN_BORDER
        if sname in SEV_FILLS:
            ws1.cell(row=sev_row, column=1).fill = SEV_FILLS[sname]
            ws1.cell(row=sev_row, column=1).font = Font(bold=True, color="FFFFFF")

    # Top alerts table
    top_row = sev_row + 2
    ws1.cell(row=top_row, column=1, value="Top Alerts by Frequency").font = Font(bold=True, size=12)
    top_row += 1
    for ci, h in enumerate(["Alert", "Count", "% of Total"], 1):
        c = ws1.cell(row=top_row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
    for item in patterns.get("alert_frequency", [])[:10]:
        top_row += 1
        ws1.cell(row=top_row, column=1, value=item["alert"]).border = THIN_BORDER
        ws1.cell(row=top_row, column=2, value=item["count"]).border = THIN_BORDER
        pct = round(item["count"] / total * 100, 1) if total else 0
        ws1.cell(row=top_row, column=3, value=f"{pct}%").border = THIN_BORDER

    # Top services
    svc_row = top_row + 2
    ws1.cell(row=svc_row, column=1, value="Top Affected Services").font = Font(bold=True, size=12)
    svc_row += 1
    for ci, h in enumerate(["Service", "Alert Count", "% of Total"], 1):
        c = ws1.cell(row=svc_row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
    for item in patterns.get("service_frequency", [])[:10]:
        svc_row += 1
        ws1.cell(row=svc_row, column=1, value=item["service"]).border = THIN_BORDER
        ws1.cell(row=svc_row, column=2, value=item["count"]).border = THIN_BORDER
        pct = round(item["count"] / total * 100, 1) if total else 0
        ws1.cell(row=svc_row, column=3, value=f"{pct}%").border = THIN_BORDER

    # Auto-width columns
    for col in range(1, 7):
        ws1.column_dimensions[get_column_letter(col)].width = 28

    # ── Severity Pie Chart in Excel ──
    chart_start = 5
    chart_end = 5 + len(sev)
    if len(sev) > 0:
        # Write hidden data for chart
        ws1.cell(row=1, column=8, value="Sev")
        ws1.cell(row=1, column=9, value="Count")
        for idx, (sname, scount) in enumerate(sev.items(), 2):
            ws1.cell(row=idx, column=8, value=sname.upper())
            ws1.cell(row=idx, column=9, value=scount)

        pie = XlPieChart()
        pie.title = "Severity Distribution"
        pie.style = 10
        labels_ref = Reference(ws1, min_col=8, min_row=2, max_row=1 + len(sev))
        data_ref = Reference(ws1, min_col=9, min_row=1, max_row=1 + len(sev))
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(labels_ref)
        pie.width = 14
        pie.height = 10
        ws1.add_chart(pie, "D5")

    # ── Sheet 2: Detailed Alerts ────────────────────────
    ws2 = wb.create_sheet("Detailed Alerts")
    ws2.sheet_properties.tabColor = "F59E0B"

    detail_cols = ["timestamp", "service", "alert", "severity", "host"]
    extra_cols = [k for k in (events[0].keys() if events else []) if k not in detail_cols]
    all_cols = detail_cols + extra_cols

    for ci, h in enumerate(all_cols, 1):
        c = ws2.cell(row=1, column=ci, value=h.replace("_", " ").title())
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER

    for ri, ev in enumerate(events, 2):
        for ci, col in enumerate(all_cols, 1):
            val = ev.get(col, "")
            if val is None:
                val = ""
            cell = ws2.cell(row=ri, column=ci, value=str(val))
            cell.border = THIN_BORDER
            # Severity colour-coding
            if col == "severity" and str(val).lower() in SEV_FILLS:
                cell.fill = SEV_FILLS[str(val).lower()]
                cell.font = Font(bold=True, color="FFFFFF")

    for col in range(1, len(all_cols) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 22
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(all_cols))}{len(events) + 1}"

    # ── Sheet 3: Root Cause Analysis ────────────────────
    ws3 = wb.create_sheet("Root Cause Analysis")
    ws3.sheet_properties.tabColor = "EF4444"

    clusters = analysis_result.get("clusters", [])
    rca_headers = ["Service", "Alert Type", "Severity", "Occurrences", "Affected Hosts", "AI Recommendation"]
    for ci, h in enumerate(rca_headers, 1):
        c = ws3.cell(row=1, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER

    suggestions = analysis_result.get("suggestions", [])
    suggestion_map = {}
    for s in suggestions:
        key = s.get("title", "")[:30]
        suggestion_map[key] = s.get("action", "Review logs and service health")

    for ri, cl in enumerate(clusters, 2):
        ws3.cell(row=ri, column=1, value=cl.get("service", "")).border = THIN_BORDER
        ws3.cell(row=ri, column=2, value=cl.get("alert_type", "")).border = THIN_BORDER
        sev_cell = ws3.cell(row=ri, column=3, value=cl.get("severity", "info").upper())
        sev_cell.border = THIN_BORDER
        if cl.get("severity") in SEV_FILLS:
            sev_cell.fill = SEV_FILLS[cl["severity"]]
            sev_cell.font = Font(bold=True, color="FFFFFF")
        ws3.cell(row=ri, column=4, value=cl.get("count", 0)).border = THIN_BORDER
        ws3.cell(row=ri, column=5, value=", ".join(cl.get("hosts", [])[:5])).border = THIN_BORDER
        # Match suggestion
        rec = "Review service logs, check resource utilization, and investigate recent changes"
        for skey, sval in suggestion_map.items():
            if cl.get("service", "").lower() in skey.lower() or cl.get("alert_type", "")[:20].lower() in skey.lower():
                rec = sval
                break
        ws3.cell(row=ri, column=6, value=rec).border = THIN_BORDER

    for col in range(1, 7):
        ws3.column_dimensions[get_column_letter(col)].width = 30

    # AI analysis text in a 4th sheet
    ai_text = analysis_result.get("ai_analysis", "")
    if ai_text:
        ws4 = wb.create_sheet("AI Analysis")
        ws4.sheet_properties.tabColor = "8B5CF6"
        ws4.merge_cells("A1:A1")
        ws4["A1"] = "AI Root Cause Analysis"
        ws4["A1"].font = Font(bold=True, size=14)
        for li, line in enumerate(str(ai_text).split("\n"), 3):
            ws4.cell(row=li, column=1, value=line)
        ws4.column_dimensions["A"].width = 120

    # Footer
    if brand.get("footer"):
        for ws in wb.worksheets:
            ws.oddFooter.center.text = brand["footer"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════
#  CHART GENERATION (Enhanced for PDF)
# ═══════════════════════════════════════════════════════════

def _dark_style(fig, ax):
    fig.patch.set_facecolor("#0F172A")
    ax.set_facecolor("#0F172A")
    for sp in ax.spines.values():
        sp.set_color("#334155")
    ax.tick_params(colors="#94A3B8", labelsize=7)
    ax.xaxis.label.set_color("#94A3B8")
    ax.yaxis.label.set_color("#94A3B8")


def _make_severity_pie(sev_dist: Dict[str, int]) -> io.BytesIO:
    fig, ax = plt.subplots(figsize=(4.2, 3.4))
    _dark_style(fig, ax)
    labels = [k.upper() for k in sev_dist.keys()]
    sizes = list(sev_dist.values())
    col_map = {"CRITICAL": "#EF4444", "WARNING": "#F59E0B", "INFO": "#06B6D4"}
    chart_colors = [col_map.get(l, "#8B5CF6") for l in labels]
    explode = [0.05] * len(labels)
    wedges, texts, autotexts = ax.pie(
        sizes, labels=labels, autopct="%1.1f%%", startangle=90, colors=chart_colors,
        textprops={"color": "white", "fontsize": 8, "fontweight": "bold"},
        explode=explode, shadow=False, pctdistance=0.75,
        wedgeprops={"edgecolor": "#0F172A", "linewidth": 2},
    )
    for at in autotexts:
        at.set_fontsize(7)
        at.set_color("white")
    centre = plt.Circle((0, 0), 0.5, fc="#0F172A")
    ax.add_artist(centre)
    total = sum(sizes)
    ax.text(0, 0.06, str(total), ha="center", va="center", fontsize=16, fontweight="bold", color="white")
    ax.text(0, -0.14, "TOTAL", ha="center", va="center", fontsize=6, color="#94A3B8")
    ax.set_title("Severity Distribution", color="white", fontsize=10, pad=12, fontweight="bold")
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)
    buf.seek(0)
    return buf


def _make_service_bar(service_freq: List[Dict], total: int) -> io.BytesIO:
    fig, ax = plt.subplots(figsize=(5.2, 3.4))
    _dark_style(fig, ax)
    items = service_freq[:8]
    names = [i["service"][:22] for i in items]
    counts = [i["count"] for i in items]
    pcts = [c / total * 100 if total else 0 for c in counts]
    bars = ax.barh(names[::-1], counts[::-1], color="#00E0FF", edgecolor="#0F172A", height=0.55)
    for bar, pct in zip(bars, pcts[::-1]):
        w = bar.get_width()
        ax.text(w + max(counts) * 0.02, bar.get_y() + bar.get_height() / 2,
                f"{pct:.0f}%", ha="left", va="center", color="#94A3B8", fontsize=7)
    ax.set_title("Top Affected Services", color="white", fontsize=10, pad=12, fontweight="bold")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.xaxis.set_major_locator(mticker.MaxNLocator(integer=True))
    ax.set_xlabel("Alert Count", fontsize=8)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)
    buf.seek(0)
    return buf


def _make_alert_bar(alert_freq: List[Dict]) -> io.BytesIO:
    fig, ax = plt.subplots(figsize=(5.2, 3.4))
    _dark_style(fig, ax)
    items = alert_freq[:8]
    names = [i["alert"][:26] for i in items]
    counts = [i["count"] for i in items]
    colors_list = ["#EF4444" if c > max(counts) * 0.7 else "#F59E0B" if c > max(counts) * 0.3 else "#06B6D4"
                   for c in counts]
    bars = ax.barh(names[::-1], counts[::-1], color=colors_list[::-1], edgecolor="#0F172A", height=0.55)
    for bar in bars:
        w = bar.get_width()
        ax.text(w + max(counts) * 0.02, bar.get_y() + bar.get_height() / 2,
                str(int(w)), ha="left", va="center", color="#94A3B8", fontsize=7)
    ax.set_title("Top Alerts by Frequency", color="white", fontsize=10, pad=12, fontweight="bold")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.xaxis.set_major_locator(mticker.MaxNLocator(integer=True))
    ax.set_xlabel("Occurrences", fontsize=8)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)
    buf.seek(0)
    return buf


def _make_severity_heatmap(clusters: List[Dict]) -> io.BytesIO:
    """Service x Severity heatmap."""
    from collections import defaultdict
    grid = defaultdict(lambda: defaultdict(int))
    for cl in clusters:
        svc = cl.get("service", "unknown")[:20]
        sev = cl.get("severity", "info")
        grid[svc][sev] += cl.get("count", 1)

    services = list(grid.keys())[:10]
    sevs = ["critical", "warning", "info"]
    matrix = [[grid[svc][s] for s in sevs] for svc in services]
    if not matrix:
        return None

    fig, ax = plt.subplots(figsize=(4.2, 3.4))
    _dark_style(fig, ax)
    import numpy as np
    data = np.array(matrix)
    im = ax.imshow(data, cmap="YlOrRd", aspect="auto")
    ax.set_xticks(range(len(sevs)))
    ax.set_xticklabels([s.upper() for s in sevs], fontsize=7)
    ax.set_yticks(range(len(services)))
    ax.set_yticklabels(services, fontsize=7)
    for i in range(len(services)):
        for j in range(len(sevs)):
            val = data[i, j]
            if val > 0:
                ax.text(j, i, str(int(val)), ha="center", va="center",
                        color="white" if val > data.max() * 0.5 else "#0F172A", fontsize=7, fontweight="bold")
    ax.set_title("Service x Severity Heatmap", color="white", fontsize=10, pad=12, fontweight="bold")
    fig.colorbar(im, ax=ax, shrink=0.8, label="Count")
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)
    buf.seek(0)
    return buf


def _make_timeline_chart(events: List[Dict]) -> io.BytesIO:
    """Alert timeline showing event density over time."""
    from collections import Counter
    from datetime import datetime as dt

    timestamps = []
    for ev in events:
        ts = ev.get("timestamp", "")
        if ts:
            try:
                if isinstance(ts, str):
                    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
                        try:
                            timestamps.append(dt.strptime(ts[:19], fmt))
                            break
                        except ValueError:
                            continue
            except Exception:
                continue

    if len(timestamps) < 2:
        return None

    timestamps.sort()
    # Group by hour
    hour_counts = Counter()
    sev_hour = {"critical": Counter(), "warning": Counter(), "info": Counter()}
    for i, ev in enumerate(events):
        ts = ev.get("timestamp", "")
        sev = ev.get("severity", "info").lower()
        if ts and i < len(timestamps):
            h = timestamps[i].strftime("%m/%d %H:00")
            hour_counts[h] += 1
            if sev in sev_hour:
                sev_hour[sev][h] += 1

    hours = sorted(hour_counts.keys())[-24:]  # Last 24 time buckets

    fig, ax = plt.subplots(figsize=(6, 3.4))
    _dark_style(fig, ax)

    x = range(len(hours))
    bottom_vals = [0] * len(hours)

    for sev_name, color in [("critical", "#EF4444"), ("warning", "#F59E0B"), ("info", "#06B6D4")]:
        vals = [sev_hour[sev_name].get(h, 0) for h in hours]
        ax.bar(x, vals, bottom=bottom_vals, color=color, width=0.7, label=sev_name.upper(), edgecolor="#0F172A")
        bottom_vals = [b + v for b, v in zip(bottom_vals, vals)]

    ax.set_xticks(x)
    ax.set_xticklabels(hours, rotation=45, ha="right", fontsize=5)
    ax.set_title("Alert Timeline", color="white", fontsize=10, pad=12, fontweight="bold")
    ax.set_ylabel("Alerts", fontsize=8)
    ax.legend(fontsize=6, loc="upper right", facecolor="#1E293B", edgecolor="#334155", labelcolor="white")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.yaxis.set_major_locator(mticker.MaxNLocator(integer=True))

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════
#  PDF GENERATION
# ═══════════════════════════════════════════════════════════

def generate_pdf_report(
    analysis_result: Dict[str, Any],
    events: List[Dict],
    branding: Optional[Dict] = None,
) -> io.BytesIO:
    """Generate professional PDF report with charts."""

    buf = io.BytesIO()
    brand = branding or {}
    report_title = brand.get("title", "FalconOps AI - Event Analysis Report")

    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=20 * mm, bottomMargin=20 * mm,
        leftMargin=15 * mm, rightMargin=15 * mm,
        title=report_title,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle("FTitle", parent=styles["Title"], fontSize=20, textColor=RL_DARK, spaceAfter=6))
    styles.add(ParagraphStyle("FH2", parent=styles["Heading2"], fontSize=14, textColor=RL_DARK, spaceBefore=16, spaceAfter=8))
    styles.add(ParagraphStyle("FBody", parent=styles["Normal"], fontSize=9, leading=13, textColor=colors.HexColor("#374151")))
    styles.add(ParagraphStyle("FSmall", parent=styles["Normal"], fontSize=8, leading=10, textColor=colors.HexColor("#6B7280")))
    styles.add(ParagraphStyle("FCenter", parent=styles["Normal"], fontSize=9, alignment=TA_CENTER))
    styles.add(ParagraphStyle("FFooter", parent=styles["Normal"], fontSize=7, alignment=TA_CENTER, textColor=colors.HexColor("#9CA3AF")))

    story = []
    summary = analysis_result.get("summary", {})
    patterns = analysis_result.get("patterns", {})
    sev = patterns.get("severity_distribution", {})
    total = summary.get("total_events", 0)

    # ── Title ──
    story.append(Paragraph(report_title, styles["FTitle"]))
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    sub_parts = [f"Generated: {ts}"]
    if brand.get("company"):
        sub_parts.append(f"Prepared for: {brand['company']}")
    story.append(Paragraph(" | ".join(sub_parts), styles["FSmall"]))
    story.append(Spacer(1, 10))
    story.append(HRFlowable(width="100%", thickness=1, color=RL_ACCENT))
    story.append(Spacer(1, 12))

    # ── Executive Summary ──
    story.append(Paragraph("1. Executive Summary", styles["FH2"]))

    kpi_data = [
        ["Metric", "Value"],
        ["Total Events", str(total)],
        ["Critical Alerts", str(sev.get("critical", 0))],
        ["Warning Alerts", str(sev.get("warning", 0))],
        ["Info Alerts", str(sev.get("info", 0))],
        ["Health Score", f"{summary.get('health_score', 0)}%"],
        ["Services Affected", str(summary.get("services_affected", 0))],
        ["Hosts Affected", str(summary.get("hosts_affected", 0))],
        ["Unique Alert Types", str(patterns.get("unique_alerts", 0))],
    ]
    t = Table(kpi_data, colWidths=[3.5 * inch, 2 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), RL_DARK),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
    ]))
    story.append(t)
    story.append(Spacer(1, 14))

    # ── Charts ──
    story.append(Paragraph("2. Visual Analysis", styles["FH2"]))

    chart_row = []
    if sev:
        pie_buf = _make_severity_pie(sev)
        chart_row.append(RLImage(pie_buf, width=3.2 * inch, height=2.6 * inch))
    service_freq = patterns.get("service_frequency", [])
    if service_freq:
        sbar_buf = _make_service_bar(service_freq, total)
        chart_row.append(RLImage(sbar_buf, width=3.8 * inch, height=2.6 * inch))

    if chart_row:
        ct = Table([chart_row], colWidths=[3.4 * inch, 4 * inch])
        ct.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
        story.append(ct)
        story.append(Spacer(1, 8))

    alert_freq = patterns.get("alert_frequency", [])
    if alert_freq:
        abar_buf = _make_alert_bar(alert_freq)
        story.append(RLImage(abar_buf, width=5 * inch, height=2.6 * inch))
        story.append(Spacer(1, 8))

    # ── Heatmap + Timeline row ──
    row2 = []
    clusters = analysis_result.get("clusters", [])
    if clusters:
        hm_buf = _make_severity_heatmap(clusters)
        if hm_buf:
            row2.append(RLImage(hm_buf, width=3.4 * inch, height=2.6 * inch))
    if events:
        tl_buf = _make_timeline_chart(events)
        if tl_buf:
            row2.append(RLImage(tl_buf, width=4 * inch, height=2.6 * inch))
    if row2:
        widths = [3.6 * inch, 4.2 * inch] if len(row2) == 2 else [7 * inch]
        ct2 = Table([row2], colWidths=widths)
        ct2.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
        story.append(ct2)
        story.append(Spacer(1, 10))

    # ── Top Services Table ──
    if service_freq:
        story.append(Paragraph("3. Top Affected Services", styles["FH2"]))
        svc_data = [["Service", "Alert Count", "% of Total"]]
        for item in service_freq[:10]:
            pct = round(item["count"] / total * 100, 1) if total else 0
            svc_data.append([item["service"], str(item["count"]), f"{pct}%"])
        st = Table(svc_data, colWidths=[3.5 * inch, 1.5 * inch, 1.5 * inch])
        st.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), RL_DARK),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ]))
        story.append(st)
        story.append(Spacer(1, 10))

    # ── Root Cause Analysis ──
    story.append(PageBreak())
    clusters = analysis_result.get("clusters", [])
    if clusters:
        story.append(Paragraph("4. Root Cause Analysis", styles["FH2"]))
        suggestions = analysis_result.get("suggestions", [])
        suggestion_map = {}
        for s in suggestions:
            key = s.get("title", "")[:30]
            suggestion_map[key] = s.get("action", "")

        rca_data = [["Service", "Alert Type", "Severity", "Count", "Recommendation"]]
        for cl in clusters[:15]:
            rec = "Review service logs and investigate"
            for skey, sval in suggestion_map.items():
                if cl.get("service", "").lower() in skey.lower():
                    rec = sval[:80]
                    break
            rca_data.append([
                cl.get("service", "")[:25],
                cl.get("alert_type", "")[:30],
                cl.get("severity", "info").upper(),
                str(cl.get("count", 0)),
                rec,
            ])
        rt = Table(rca_data, colWidths=[1.3 * inch, 1.6 * inch, 0.7 * inch, 0.5 * inch, 3 * inch])
        rt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), RL_DARK),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
            ("ALIGN", (2, 0), (3, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        # Colour severity cells
        for ri, cl in enumerate(clusters[:15], 1):
            sev_name = cl.get("severity", "info")
            if sev_name in RL_SEV:
                rt.setStyle(TableStyle([
                    ("BACKGROUND", (2, ri), (2, ri), RL_SEV[sev_name]),
                    ("TEXTCOLOR", (2, ri), (2, ri), colors.white),
                ]))
        story.append(rt)
        story.append(Spacer(1, 10))

    # ── AI Analysis Text ──
    ai_text = analysis_result.get("ai_analysis", "")
    if ai_text:
        story.append(Paragraph("5. AI Root Cause Analysis", styles["FH2"]))
        for line in str(ai_text).split("\n"):
            line = line.strip()
            if not line:
                story.append(Spacer(1, 4))
            elif line.startswith("###"):
                story.append(Paragraph(line.replace("#", "").strip(), styles["Heading4"]))
            elif line.startswith("##"):
                story.append(Paragraph(line.replace("#", "").strip(), styles["Heading3"]))
            elif line.startswith("#"):
                story.append(Paragraph(line.replace("#", "").strip(), styles["FH2"]))
            elif line.startswith("|"):
                continue  # skip markdown tables
            elif line.startswith("```"):
                continue
            elif line.startswith("- ") or line.startswith("* "):
                story.append(Paragraph(f"&bull; {line[2:]}", styles["FBody"]))
            else:
                # Escape XML special chars
                safe_line = line.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                story.append(Paragraph(safe_line, styles["FBody"]))

    # ── Suggestions ──
    suggestions = analysis_result.get("suggestions", [])
    if suggestions:
        story.append(Spacer(1, 10))
        story.append(Paragraph("6. AI Recommendations", styles["FH2"]))
        for i, s in enumerate(suggestions[:8], 1):
            pri = s.get("priority", "medium").upper()
            title = s.get("title", "")
            safe_title = title.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            desc = (s.get("description", "") or "")[:200]
            safe_desc = desc.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            story.append(Paragraph(f"<b>[{pri}] {i}. {safe_title}</b>", styles["FBody"]))
            story.append(Paragraph(safe_desc, styles["FSmall"]))
            story.append(Spacer(1, 6))

    # ── Footer ──
    story.append(Spacer(1, 20))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#D1D5DB")))
    footer_text = brand.get("footer", "FalconOps AI - Enterprise AIOps Platform | Confidential")
    story.append(Paragraph(footer_text, styles["FFooter"]))

    doc.build(story)
    buf.seek(0)
    return buf
